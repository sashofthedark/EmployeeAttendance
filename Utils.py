from Errors import NotInDB,InDB


class Employee():
    def __init__(self,id,name,phone,age):
        self.id = id
        self.name = name
        self.phone = phone
        self.age = age

class EmployeeList():
    def CreateEmployeeList(self):
        try:
            self.list
        except AttributeError:
            self.list = []
    
    def check(self,id,name,phone,age):
        #does this function "see" the self.list?
        #checks whether the data for the employee is valid and he's not already in the employee list
        self.CreateEmployeeList()
        a = [(ind,value) for (ind,value) in enumerate(self.list) if value.id == id]
        if a!=[]:
            raise InDB('Employee is already in database!')
        elif type(id)!= int or len(str(id)) != 4:
            raise ValueError('Invalid ID')
        elif type(name)!=str:
            raise ValueError('Invalid name')
        elif type(phone)!=int or len(str(phone))!=10:
            raise ValueError('Invalid phone number')
        elif type(age)!=int or int(age)<18 or int(age)>67:
            raise ValueError('Invalid age')
        else:
            return True
       
    def add(self,id,name,phone,age):
        self.CreateEmployeeList()
        try:
            self.check(id,name,phone,age)
        except InDB as NDB:
            print(f'Error message was {NDB}')
        except ValueError as ValErr:
            print(f'Error message was {ValErr}')
            return False
        else:
            self.list.append(Employee(id,name,phone,age))
            return True

    def delete(self,employeeID):
        Deleted = False
        if type(employeeID)!= int or len(str(employeeID))!=4:
            raise ValueError('Invalid id, should be 4 digits long')
        else:
            a = [(ind,value) for (ind,value) in enumerate(self.list) if value.id == employeeID]
            if len(a) == 0:
                raise ValueError('ID doesn\'t exist')
            else:
                #print(f'ID for deletion {a[0][1].id}')
                del self.list[a[0][0]]
                Deleted = True
        return Deleted 
        #we return a binary value telling the user whether the employee was removed from the database        

    def addFromCSV(self,filepath):
        import csv
        import os

        self.CreateEmployeeList()

        if filepath[-4:]!='.csv':
            raise TypeError('Wrong file type')
        else:
            with open(filepath) as csv_file:
                csv_reader = csv.reader(csv_file,delimiter=',')
                if os.stat(filepath).st_size == 0:
                    print('File is empty, nothing to import')
                else:
                    for row in csv_reader:
                        try:
                            idr = int(row[0])
                            namer = str(row[1])
                            phoner = int(row[2])
                            ager = int(row[3])
                            self.check(idr,namer,phoner,ager)
                        except InDB:
                            pass
                        except Exception as Exc:
                            raise
                        else:
                            self.add(idr,namer,phoner,ager)

    def addFromXLSX(self,filepath):
        import os
        import openpyxl

        self.CreateEmployeeList()

        if filepath[-5:]!='.xlsx':
            raise TypeError('Wrong file type')
        else:
            wb_obj = openpyxl.load_workbook(filepath)
            sheet_obj = wb_obj.active 
            for rows in range(1,sheet_obj.max_row+1):
                try:
                    idc = int(sheet_obj.cell(row = rows,column = 1).value)
                    namec = str(sheet_obj.cell(row = rows,column = 2).value)
                    phonec = int(sheet_obj.cell(row = rows,column = 3).value)
                    agec = int(sheet_obj.cell(row = rows,column = 4).value)

                    self.check(idc,namec,phonec,agec)
                except InDB as NDB:
                    print(f'The following exception occurred {NDB} in row {rows}')
                    raise InDB
                except Exception as Exc:
                    print(f'The following exception occurred {Exc} in row {rows}')
                    raise ValueError
                else:
                    self.add(idc,namec,phonec,agec)

    def delFromCSV(self,filepath):
        import csv
        import os

        self.CreateEmployeeList()

        if filepath[-4:]!='.csv':
            raise TypeError('Wrong file type')
        else:
            with open(filepath) as csv_file:
                csv_reader = csv.reader(csv_file,delimiter=',')
                if os.stat(filepath).st_size == 0:
                    print('File is empty, nothing to delete')
                else:
                    row_cnt = 1
                    for row in csv_reader:
                        try:
                            idr = int(row[0])
                            namer = str(row[1])
                            phoner = int(row[2])
                            ager = int(row[3])
                            self.check(idr,namer,phoner,ager)
                        except InDB:
                            self.delete(idr)
                            row_cnt+=1
                        except ValueError as Exc:
                            print(f'The following error occurred: {Exc} in row {str(row_cnt)}')
                            raise

    def delFromXLSX(self,filepath):
        import os
        import openpyxl

        self.CreateEmployeeList()

        if filepath[-5:]!='.xlsx':
            raise TypeError('Wrong file type')
        else:
            wb_obj = openpyxl.load_workbook(filepath)
            sheet_obj = wb_obj.active 
            row_cnt = 1
            for rows in range(1,sheet_obj.max_row+1):
                try:
                    idc = int(sheet_obj.cell(row = rows,column = 1).value)
                    namec = str(sheet_obj.cell(row = rows,column = 2).value)
                    phonec = int(sheet_obj.cell(row = rows,column = 3).value)
                    agec = int(sheet_obj.cell(row = rows,column = 4).value)

                    self.check(idc,namec,phonec,agec)
                except InDB:
                    self.delete(idc)
                    row_cnt+=1
                except Exception as Exc:
                    print(f'The following exception occurred {Exc} in row {rows}')
                    raise ValueError
                    


